import discord
import logging
import openpyxl
import os
import asyncio
import tempfile
from typing import List, Any
from pathlib import Path
from discord import app_commands
from discord.ext import commands
from datetime import datetime
from src.config import TIMEZONE
from openpyxl.styles import Font, Alignment

logger = logging.getLogger("PontoBot.Report")


class ReportCog(commands.Cog):
    def __init__(self, bot):
        self.bot = bot
        self.db = bot.db

    @app_commands.command(
        name="relatorio", description="Gerar relatório de horas em Excel"
    )
    @app_commands.describe(
        usuario="Usuário específico para o relatório (deixe vazio para o seu)"
    )
    async def relatorio(
        self, interaction: discord.Interaction, usuario: discord.Member = None
    ):
        await interaction.response.defer(ephemeral=True)

        target = usuario or interaction.user
        guild_id = interaction.guild_id

        # Check de permissão: apenas o próprio usuário ou admin/moderador pode ver o relatório
        if target.id != interaction.user.id:
            if (
                not interaction.user.guild_permissions.administrator
                and not interaction.user.guild_permissions.manage_guild
            ):
                # Opcional: checar cargo de moderador configurado no banco aqui
                await interaction.followup.send(
                    "❌ Você não tem permissão para ver o relatório de outros usuários.",
                    ephemeral=True,
                )
                return

        try:
            registros = await self.db.get_user_records(target.id, guild_id)
        except Exception as e:
            logger.error(
                f"Erro ao buscar registros para relatório do usuário {target.id}: {e}"
            )
            await interaction.followup.send(
                "❌ Erro ao acessar o banco de dados.", ephemeral=True
            )
            return

        if not registros:
            await interaction.followup.send(
                "❌ Nenhum registro encontrado para este usuário.", ephemeral=True
            )
            return

        filename: str = ""
        try:
            # Gerar Excel em uma thread separada
            try:
                loop = asyncio.get_running_loop()
                filename = await loop.run_in_executor(
                    None, self._generate_excel, target, registros
                )
            except Exception:
                logger.exception(f"Erro ao gerar arquivo Excel para {target.id}")
                await interaction.followup.send(
                    "❌ Erro ao gerar o relatório.", ephemeral=True
                )
                return

            embed = discord.Embed(
                title="📊 Relatório Gerado",
                description=f"O histórico de pontos de **{target.display_name}** foi processado.",
                color=discord.Color.purple(),
                timestamp=datetime.now(TIMEZONE),
            )
            embed.set_thumbnail(url=target.display_avatar.url)
            embed.add_field(
                name="Total de Registros", value=str(len(registros)), inline=True
            )

            # Verificação defensiva do arquivo
            if not filename or not os.path.exists(filename):
                logger.error(f"Arquivo de relatório não encontrado: {filename}")
                await interaction.followup.send(
                    "❌ Erro interno: O arquivo de relatório não foi criado.", ephemeral=True
                )
                return

            try:
                await interaction.followup.send(
                    embed=embed, file=discord.File(filename), ephemeral=True
                )
            except Exception:
                logger.exception("Erro ao enviar arquivo de relatório")
                await interaction.followup.send(
                    "❌ Erro ao enviar o arquivo de relatório.", ephemeral=True
                )
        finally:
            # Limpeza garantida
            if filename and os.path.exists(filename):
                try:
                    os.remove(filename)
                except Exception as e:
                    logger.error(f"Erro ao deletar arquivo temporário {filename}: {e}")

    def _generate_excel(self, target: discord.Member, registros: List[Any]) -> str:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Relatório de Ponto"

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = openpyxl.styles.PatternFill(
            start_color="5865F2", end_color="5865F2", fill_type="solid"
        )
        center_align = Alignment(horizontal="center")

        # Cabeçalho
        headers = ["Data/Hora", "Tipo", "Duração"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=text)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # Dados
        for idx, row in enumerate(registros, 2):
            try:
                # Safe keys
                ts_val = row.get("timestamp") or str(datetime.now(TIMEZONE).isoformat())
                dt = datetime.fromisoformat(ts_val)
                # Garantir TZ awareness
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=TIMEZONE)

                # Format using local time
                ws.cell(row=idx, column=1, value=dt.strftime("%d/%m/%Y %H:%M:%S"))
            except (ValueError, TypeError) as e:
                logger.warning(
                    f"Timestamp inválido no registro {row.get('id', '?')}: {e}"
                )
                ws.cell(row=idx, column=1, value=str(row.get("timestamp", "-")))

            tipo = str(row.get("tipo", "")).capitalize()
            ws.cell(row=idx, column=2, value=tipo)

            duracao = row.get("duracao_segundos")
            if tipo.lower() == "saida" and isinstance(duracao, (int, float)):
                horas = int(duracao) // 3600
                minutos = (int(duracao) % 3600) // 60
                ws.cell(row=idx, column=3, value=f"{horas}h {minutos}min")
            else:
                ws.cell(row=idx, column=3, value="-")

        # Ajustar largura das colunas
        try:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        val = str(cell.value) if cell.value is not None else ""
                        if len(val) > max_length:
                            max_length = len(val)
                    except (ValueError, TypeError, AttributeError):
                        pass
                adjusted_width = max_length + 2
                ws.column_dimensions[column].width = adjusted_width
        except Exception as e:
            logger.warning(f"Erro ao ajustar larguras de coluna: {e}")

        # Gerar nome de arquivo seguro usando tempfile
        suffix = f"_{datetime.now(TIMEZONE).strftime('%Y%m%d_%H%M%S')}.xlsx"
        # Criar arquivo temporário seguro
        fd, path = tempfile.mkstemp(suffix=suffix, prefix="relatorio_")
        os.close(fd)  # Fechar o file descriptor pois wb.save abrirá o arquivo

        wb.save(path)
        return path


async def setup(bot):
    await bot.add_cog(ReportCog(bot))
