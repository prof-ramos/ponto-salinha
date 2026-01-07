import discord
from discord.ext import commands
from discord import app_commands
from database import Database
from datetime import datetime, timedelta
import os
import openpyxl
from openpyxl.styles import Font, Alignment
import pytz
from functools import lru_cache

intents = discord.Intents.default()
intents.members = True
intents.message_content = True

bot = commands.Bot(command_prefix="!", intents=intents)

# Helpers for Timezone
@lru_cache(maxsize=1024)
def get_guild_timezone(guild_id: int):
    """Retrieve timezone string from DB or default to Sao_Paulo"""
    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT timezone FROM config WHERE guild_id = ?', (guild_id,))
    res = cursor.fetchone()
    conn.close()
    if res and res[0]:
        return pytz.timezone(res[0])
    return pytz.timezone('America/Sao_Paulo')

def get_now(guild_id: int):
    """Get current time aware of guild timezone"""
    tz = get_guild_timezone(guild_id)
    return datetime.now(tz)

# Role Verification Check
async def check_authorized_role(interaction: discord.Interaction) -> bool:
    """
    Check if user has the authorized role OR is an administrator.
    Returns True if authorized, else sends ephemeral error and returns False.
    """
    # Check if admin
    if interaction.user.guild_permissions.administrator:
        return True

    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT cargo_autorizado_id FROM config WHERE guild_id = ?', (interaction.guild_id,))
    res = cursor.fetchone()
    conn.close()

    # If no role configured, only admin (already checked)
    if not res or not res[0]:
        await interaction.response.send_message("❌ Nenhum cargo autorizado configurado e você não é administrador.", ephemeral=True)
        return False

    role_id = res[0]
    user_roles = [r.id for r in interaction.user.roles]

    if role_id in user_roles:
        return True

    await interaction.response.send_message("❌ Você não tem permissão para usar este comando.", ephemeral=True)
    return False


@bot.event
async def on_ready():
    print(f'Bot conectado como {bot.user}')
    # Initialize DB on startup to ensure tables exist
    Database().init_db()
    try:
        synced = await bot.tree.sync()
        print(f'Comandos sincronizados! ({len(synced)} comandos)')
    except Exception as e:
        print(f"Erro ao sincronizar comandos: {e}")

@bot.tree.command(name="config", description="Configurar o bot")
@app_commands.describe(
    canal_log="Canal para enviar logs de ponto",
    cargo="Cargo autorizado a usar comandos",
    fuso_horario="Fuso horário (Ex: America/Sao_Paulo)"
)
async def config(
    interaction: discord.Interaction,
    canal_log: discord.TextChannel = None,
    cargo: discord.Role = None,
    fuso_horario: str = None
):
    if not interaction.user.guild_permissions.administrator:
        await interaction.response.send_message("❌ Apenas administradores podem configurar o bot.", ephemeral=True)
        return

    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Get current values first to handle partial updates
    cursor.execute('SELECT log_channel_id, cargo_autorizado_id, timezone FROM config WHERE guild_id = ?', (interaction.guild_id,))
    current = cursor.fetchone()

    new_log_id = canal_log.id if canal_log else (current[0] if current else None)
    new_role_id = cargo.id if cargo else (current[1] if current else None)
    new_timezone = fuso_horario if fuso_horario else (current[2] if current else 'America/Sao_Paulo')

    # Validate timezone
    try:
        pytz.timezone(new_timezone)
    except pytz.UnknownTimeZoneError:
        conn.close()
        await interaction.response.send_message(f"❌ Fuso horário inválido: {new_timezone}. Use formato 'Continente/Cidade' (Ex: America/Sao_Paulo)", ephemeral=True)
        return

    cursor.execute('''
        INSERT OR REPLACE INTO config (guild_id, log_channel_id, cargo_autorizado_id, timezone)
        VALUES (?, ?, ?, ?)
    ''', (interaction.guild_id, new_log_id, new_role_id, new_timezone))
    
    conn.commit()
    conn.close()

    # Clear cache for this guild (since we can't easily clear just one entry in lru_cache, we clear all)
    get_guild_timezone.cache_clear()
    
    msg_parts = []
    if canal_log: msg_parts.append(f"Logs: {canal_log.mention}")
    if cargo: msg_parts.append(f"Cargo: {cargo.mention}")
    if fuso_horario: msg_parts.append(f"Fuso: {fuso_horario}")

    await interaction.response.send_message(
        f"✅ Configuração atualizada!\n" + "\n".join(msg_parts),
        ephemeral=True
    )

@bot.tree.command(name="ponto", description="Registrar entrada ou saída")
async def ponto(interaction: discord.Interaction):
    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Verificar status atual
    cursor.execute('''
        SELECT status, timestamp_entrada FROM status_ponto
        WHERE user_id = ? AND guild_id = ?
    ''', (interaction.user.id, interaction.guild_id))
    
    result = cursor.fetchone()

    # Use timezone-aware current time
    now_aware = get_now(interaction.guild_id)
    timestamp_atual = now_aware.isoformat()
    
    if not result or result[0] == 'inativo':
        # ENTRADA
        cursor.execute('''
            INSERT OR REPLACE INTO status_ponto (user_id, guild_id, status, timestamp_entrada)
            VALUES (?, ?, 'ativo', ?)
        ''', (interaction.user.id, interaction.guild_id, timestamp_atual))
        
        cursor.execute('''
            INSERT INTO registros (user_id, guild_id, timestamp, tipo)
            VALUES (?, ?, ?, 'entrada')
        ''', (interaction.user.id, interaction.guild_id, timestamp_atual))
        
        tipo_msg = "entrada"
        emoji = "🟢"
    else:
        # SAÍDA
        timestamp_entrada = datetime.fromisoformat(result[1])
        # Ensure timestamp_entrada is aware if it wasn't before (handling migration/mixed data)
        if timestamp_entrada.tzinfo is None:
             timestamp_entrada = get_guild_timezone(interaction.guild_id).localize(timestamp_entrada)

        timestamp_saida = now_aware
        duracao = int((timestamp_saida - timestamp_entrada).total_seconds())
        
        cursor.execute('''
            INSERT INTO registros (user_id, guild_id, timestamp, tipo, duracao_segundos)
            VALUES (?, ?, ?, 'saida', ?)
        ''', (interaction.user.id, interaction.guild_id, timestamp_atual, duracao))
        
        cursor.execute('''
            UPDATE status_ponto SET status = 'inativo'
            WHERE user_id = ? AND guild_id = ?
        ''', (interaction.user.id, interaction.guild_id))
        
        tipo_msg = "saída"
        emoji = "🔴"
    
    conn.commit()
    
    # Buscar canal de log
    cursor.execute('SELECT log_channel_id FROM config WHERE guild_id = ?', 
                   (interaction.guild_id,))
    config_data = cursor.fetchone()
    conn.close()
    
    await interaction.response.send_message(
        f"{emoji} **Ponto de {tipo_msg}** registrado!",
        ephemeral=True
    )
    
    # Enviar para canal de log
    if config_data and config_data[0]:
        canal = interaction.guild.get_channel(config_data[0])
        if canal:
            await canal.send(
                f"{emoji} {interaction.user.mention} registrou **{tipo_msg}** às {now_aware.strftime('%H:%M:%S')}"
            )

@bot.tree.command(name="ranking", description="Ver ranking de horas trabalhadas")
@app_commands.describe(periodo="Período para calcular")
@app_commands.choices(periodo=[
    app_commands.Choice(name="Hoje", value="hoje"),
    app_commands.Choice(name="Esta Semana", value="semana"),
    app_commands.Choice(name="Este Mês", value="mes"),
    app_commands.Choice(name="Total", value="total")
])
async def ranking(interaction: discord.Interaction, periodo: str = "semana"):
    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    # Calcular data de início
    now = get_now(interaction.guild_id)
    
    if periodo == "hoje":
        data_inicio = now.replace(hour=0, minute=0, second=0).isoformat()
    elif periodo == "semana":
        data_inicio = (now - timedelta(days=7)).isoformat()
    elif periodo == "mes":
        data_inicio = (now - timedelta(days=30)).isoformat()
    else:
        data_inicio = "2000-01-01"
    
    # Buscar horas trabalhadas
    cursor.execute('''
        SELECT user_id, SUM(duracao_segundos) as total_segundos
        FROM registros
        WHERE guild_id = ? AND tipo = 'saida' AND timestamp >= ?
        GROUP BY user_id
        ORDER BY total_segundos DESC
        LIMIT 10
    ''', (interaction.guild_id, data_inicio))
    
    resultados = cursor.fetchall()
    conn.close()
    
    if not resultados:
        await interaction.response.send_message("📊 Nenhum registro encontrado.", ephemeral=True)
        return
    
    # Formatando ranking
    embed = discord.Embed(
        title=f"🏆 Ranking - {periodo.capitalize()}",
        color=discord.Color.gold()
    )
    
    for idx, (user_id, total_seg) in enumerate(resultados, 1):
        user = interaction.guild.get_member(user_id)
        if total_seg is None:
            total_seg = 0
            
        horas = total_seg // 3600
        minutos = (total_seg % 3600) // 60
        
        medalha = ["🥇", "🥈", "🥉"][idx-1] if idx <= 3 else f"{idx}º"
        nome = user.display_name if user else f"Usuário {user_id}"
        
        embed.add_field(
            name=f"{medalha} {nome}",
            value=f"⏱️ {horas}h {minutos}min",
            inline=False
        )
    
    await interaction.response.send_message(embed=embed)

@bot.tree.command(name="relatorio", description="Gerar relatório de horas")
@app_commands.describe(usuario="Usuário para relatório (deixe vazio para você)")
async def relatorio(interaction: discord.Interaction, usuario: discord.Member = None):
    # Check authorization
    if not await check_authorized_role(interaction):
        return

    await interaction.response.defer(ephemeral=True)
    
    target = usuario or interaction.user
    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    cursor.execute('''
        SELECT timestamp, tipo, duracao_segundos
        FROM registros
        WHERE user_id = ? AND guild_id = ?
        ORDER BY timestamp DESC
        LIMIT 100
    ''', (target.id, interaction.guild_id))
    
    registros = cursor.fetchall()
    conn.close()
    
    if not registros:
        await interaction.followup.send("❌ Nenhum registro encontrado.", ephemeral=True)
        return
    
    # Gerar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Ponto"
    
    # Cabeçalho
    ws['A1'] = "Data/Hora"
    ws['B1'] = "Tipo"
    ws['C1'] = "Duração"
    
    for cell in ['A1', 'B1', 'C1']:
        ws[cell].font = Font(bold=True)
        ws[cell].alignment = Alignment(horizontal='center')
    
    # Dados
    for idx, (timestamp_str, tipo, duracao) in enumerate(registros, 2):
        dt = datetime.fromisoformat(timestamp_str)
        # Assuming report dates should also follow guild timezone?
        # For excel export, we keep as is, or can convert if needed.
        # But `fromisoformat` preserves offset if present.

        ws[f'A{idx}'] = dt.strftime('%d/%m/%Y %H:%M:%S')
        ws[f'B{idx}'] = tipo.capitalize()
        
        if tipo == 'saida' and duracao:
            horas = duracao // 3600
            minutos = (duracao % 3600) // 60
            ws[f'C{idx}'] = f"{horas}h {minutos}min"
    
    filename = f"relatorio_{target.name}_{get_now(interaction.guild_id).strftime('%Y%m%d')}.xlsx"
    wb.save(filename)
    
    await interaction.followup.send(
        f"📊 Relatório de **{target.display_name}**",
        file=discord.File(filename),
        ephemeral=True
    )
    
    try:
        os.remove(filename)
    except Exception as e:
        print(f"Erro ao deletar arquivo temporário: {e}")

@bot.tree.command(name="limpar_dados", description="Limpar registros de ponto")
@app_commands.describe(periodo="Período para limpar")
@app_commands.choices(periodo=[
    app_commands.Choice(name="Último Dia", value="dia"),
    app_commands.Choice(name="Última Semana", value="semana"),
    app_commands.Choice(name="Último Mês", value="mes"),
    app_commands.Choice(name="Todos os Dados", value="total")
])
# @app_commands.checks.has_permissions(administrator=True) # Removed decorator to use custom check
async def limpar_dados(interaction: discord.Interaction, periodo: str):
    
    # Use custom check for role or admin
    if not await check_authorized_role(interaction):
        return

    db = Database()
    conn = db.get_connection()
    cursor = conn.cursor()
    
    if periodo == "total":
        cursor.execute('DELETE FROM registros WHERE guild_id = ?', 
                      (interaction.guild_id,))
        cursor.execute('DELETE FROM status_ponto WHERE guild_id = ?', 
                      (interaction.guild_id,))
    else:
        dias = {"dia": 1, "semana": 7, "mes": 30}[periodo]
        # Use timezone aware comparison? Or just UTC/Naive?
        # Since timestamps are stored as strings (isoformat), if they have offset, comparison works.
        # But `datetime.now()` needs to match the stored format's offset awareness.

        limit_date = get_now(interaction.guild_id) - timedelta(days=dias)
        data_limite = limit_date.isoformat()
        
        cursor.execute('''
            DELETE FROM registros 
            WHERE guild_id = ? AND timestamp < ?
        ''', (interaction.guild_id, data_limite))
    
    linhas = cursor.rowcount
    conn.commit()
    conn.close()
    
    await interaction.response.send_message(
        f"🗑️ **{linhas} registros** foram removidos ({periodo}).",
        ephemeral=True
    )

# Substitua pelo seu token ou use os.getenv('DISCORD_TOKEN')
if __name__ == "__main__":
    TOKEN = os.getenv('DISCORD_TOKEN')
    if not TOKEN:
        print("Erro: Token não encontrado. Defina a variável de ambiente DISCORD_TOKEN ou edite o arquivo.")
    else:
        bot.run(TOKEN)
